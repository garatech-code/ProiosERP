from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging

from apps.operaciones.models import Operacion, Client, Ship, Port, Agency, AgendaEvent, DocumentoAdjunto
from apps.usuarios.models import User
from apps.operaciones.services import get_or_create_ship_from_imo, get_or_create_port_from_name
from apps.produccion.models import OrdenFabricacion
from rest_framework.permissions import IsAuthenticated
from apps.usuarios.permissions import IsOwnerOrCreatorSenior

from .serializers import (
    OperacionSerializer, ClientSerializer, ShipSerializer,
    PortSerializer, AgencySerializer, OperacionDetalleSerializer,
    AgendaEventSerializer, DocumentoAdjuntoSerializer
)

logger = logging.getLogger(__name__)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class ShipViewSet(viewsets.ModelViewSet):
    queryset = Ship.objects.all()
    serializer_class = ShipSerializer


class PortViewSet(viewsets.ModelViewSet):
    queryset = Port.objects.all()
    serializer_class = PortSerializer


class AgencyViewSet(viewsets.ModelViewSet):
    queryset = Agency.objects.all()
    serializer_class = AgencySerializer


class OperacionViewSet(viewsets.ModelViewSet):
    serializer_class = OperacionSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrCreatorSenior]

    def get_queryset(self):
        user = self.request.user
        print(f"🔍 [BACKEND] Usuario: {user.username} (role: {user.role})")
        qs = Operacion.objects.all().select_related('cliente', 'ship', 'port', 'agency').prefetch_related('detalles', 'documentos_adjuntos')

        if user.role in [User.Role.OWNER, User.Role.CONTABLE]:
            print("✅ Owner/Contable -> todas las operaciones")
            return qs

        if user.role == User.Role.OPERADOR:
            filtered = qs.filter(Q(operadores_asignados=user) | Q(creado_por=user)).distinct()
            print(f"👤 OPERADOR -> {filtered.count()} operaciones asignadas o creadas")
            return filtered

        if getattr(User.Role, 'OPERADOR_JR', None) and user.role == User.Role.OPERADOR_JR:
            filtered = qs.filter(operadores_asignados=user).distinct()
            print(f"👦 OPERADOR JR -> {filtered.count()} operaciones asignadas")
            return filtered

        if user.role == User.Role.OPERARIO:
            filtered = qs.filter(operarios_usuarios_asignados=user).distinct()
            print(f"👷 OPERARIO -> {filtered.count()} operaciones asignadas")
            print(f"IDs: {list(filtered.values_list('id', flat=True))}")
            return filtered

        print("⚠️ Rol no reconocido -> 0 operaciones")
        return qs.none()

    def perform_create(self, serializer):
        user = self.request.user
        operation = serializer.save(creado_por=user)
        
        if user.role not in [User.Role.OWNER, User.Role.CONTABLE]:
            if user.role == User.Role.OPERADOR:
                operadores_ids = self.request.data.get('operadores_asignados', [])
                if operadores_ids:
                    operation.operadores_asignados.set(operadores_ids)
                else:
                    operation.operadores_asignados.add(user)
            elif getattr(User.Role, 'OPERADOR_JR', None) and user.role == User.Role.OPERADOR_JR:
                operation.operadores_asignados.add(user)
            elif user.role == User.Role.OPERARIO:
                operation.operarios_usuarios_asignados.add(user)
                
            operation.estado_revision = Operacion.ESTADO_REVISION_PENDING
            operation.mensaje_revision = f"Operación iniciada por {user.username} ({user.role}) el {timezone.now().strftime('%d/%m/%Y a las %H:%M')}"
            operation.save()

    @action(detail=False, methods=['get'])
    def dashboard_metrics(self, request):
        user = request.user
        if user.role not in [User.Role.OWNER, User.Role.CONTABLE]:
            return Response({'error': 'No autorizado'}, status=403)

        qs = Operacion.objects.all()

        # Operaciones
        total_ops = qs.count()
        ops_en_proceso = qs.exclude(estado__in=[
            Operacion.ESTADO_ENTREGADA,
            Operacion.ESTADO_CANCELADA
        ]).count()
        ops_finalizadas = qs.filter(estado=Operacion.ESTADO_ENTREGADA).count()
        ops_canceladas = qs.filter(estado=Operacion.ESTADO_CANCELADA).count()
        ops_revision = qs.filter(estado_revision=Operacion.ESTADO_REVISION_PENDING).count()

        ops_productos = qs.filter(tipo_operacion='productos').count()
        ops_quimicos = qs.filter(tipo_operacion='quimicos').count()
        ops_servicios = qs.filter(tipo_operacion='servicios').count()
        ops_otros = qs.filter(tipo_operacion='otros').count()

        # Usuarios
        usuarios_activos = User.objects.filter(is_active=True).count()
        usuarios_owner = User.objects.filter(is_active=True, role=User.Role.OWNER).count()
        
        # Safe handling of roles that might not exist in older migrations
        operador_jr_role = getattr(User.Role, 'OPERADOR_JR', 'OPERADOR_JR')
        usuarios_operador = User.objects.filter(is_active=True, role__in=[User.Role.OPERADOR, operador_jr_role]).count()
        usuarios_operario = User.objects.filter(is_active=True, role=User.Role.OPERARIO).count()

        # Inventario
        from apps.inventario.models import Articulo
        from django.db.models import F
        total_articulos = Articulo.objects.count()
        articulos_bajo_stock = Articulo.objects.filter(controlar_stock=True, stock_actual__lte=F('stock_minimo')).count()
        articulos_quimicos = Articulo.objects.filter(categoria='quimicos').count()
        articulos_otros = total_articulos - articulos_quimicos

        return Response({
            'operaciones': {
                'total': total_ops,
                'en_proceso': ops_en_proceso,
                'finalizadas': ops_finalizadas,
                'canceladas': ops_canceladas,
                'revision': ops_revision,
                'por_tipo': {
                    'productos': ops_productos,
                    'quimicos': ops_quimicos,
                    'servicios': ops_servicios,
                    'otros': ops_otros
                }
            },
            'usuarios': {
                'activos': usuarios_activos,
                'owner': usuarios_owner,
                'operadores': usuarios_operador,
                'operarios': usuarios_operario
            },
            'inventario': {
                'total_articulos': total_articulos,
                'bajo_stock': articulos_bajo_stock,
                'quimicos': articulos_quimicos,
                'otros': articulos_otros
            },
            # Retrocompatibilidad
            'total': total_ops,
            'en_proceso': ops_en_proceso,
            'finalizadas': ops_finalizadas,
            'usuarios_activos': usuarios_activos
        })

    @action(detail=False, methods=['get'])
    def holidays_ar(self, request):
        try:
            import holidays
            year = timezone.now().year
            ar_holidays = holidays.AR(years=[year, year+1])
            data = [{"date": str(date), "name": name} for date, name in ar_holidays.items()]
            return Response(data)
        except ImportError:
            return Response({'error': 'Librería holidays no instalada'}, status=500)

    def destroy(self, request, *args, **kwargs):
        if getattr(User.Role, 'OPERADOR_JR', None) and request.user.role == User.Role.OPERADOR_JR:
            return Response({'error': 'No tiene permisos para eliminar operaciones'}, status=403)
        return super().destroy(request, *args, **kwargs)

    def notificar_senior(self, operacion, accion):
        if operacion.creado_por and operacion.creado_por.role == User.Role.OPERADOR:
            if getattr(User.Role, 'OPERADOR_JR', None) and self.request.user.role == User.Role.OPERADOR_JR:
                from apps.usuarios.models import Notificacion
                Notificacion.objects.create(
                    usuario_destino=operacion.creado_por,
                    mensaje=f"El Junior {self.request.user.username} ha {accion} la operación OP-{operacion.id}.",
                    operacion_id=operacion.id
                )

    @action(detail=True, methods=['post'])
    def cancel_operation(self, request, pk=None):
        if getattr(User.Role, 'OPERADOR_JR', None) and request.user.role == User.Role.OPERADOR_JR:
            return Response({'error': 'No tiene permisos para anular operaciones'}, status=403)
        op = self.get_object()
        try:
            with transaction.atomic():
                op.cancel()
                op.save()
            return Response({'status': 'cancelled'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


    @action(detail=True, methods=['post'])
    def confirm_operation(self, request, pk=None):
        operacion = self.get_object()

        if operacion.estado != Operacion.ESTADO_SOLICITADA:
            return Response(
                {'error': f'La operación no está en estado inicial. Estado actual: {operacion.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            try:
                operacion.start_packing()
                operacion.save()
                self.notificar_senior(operacion, "iniciado el armado (confirmado)")
            except ValidationError as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'error': f'Error inesperado: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        serializer = self.get_serializer(operacion)
        return Response({
            'status': 'confirmed',
            'operation': serializer.data,
            'message': 'Operación confirmada y stock consumido correctamente'
        })

    @action(detail=True, methods=['post'])
    def start_coordination(self, request, pk=None):
        op = self.get_object()
        if not op.packing_list_file:
            proveedor = request.data.get('proveedor', 'PROIOS SA')
            pais_destino = request.data.get('pais_destino', 'argentina')
            from apps.operaciones.services import generar_y_guardar_packing_list
            try:
                generar_y_guardar_packing_list(op, proveedor=proveedor, pais_destino=pais_destino)
            except Exception as e:
                logger.exception("Error generando y guardando packing list automáticamente")
                return Response({'error': f'No se pudo autogenerar el Packing List: {str(e)}'}, status=400)

        with transaction.atomic():
            try:
                op.send_to_customs()
                op.save()
                self.notificar_senior(op, "avanzado a aduanas")
                return Response({'status': 'in_customs'})
            except ValidationError as e:
                return Response({'error': str(e)}, status=400)
            except Exception as e:
                return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def finalize_production(self, request, pk=None):
        op = self.get_object()
        if not op.rancho_file:
            return Response({'error': 'Debe subir el Documento Rancho antes de aprobar aduanas'}, status=400)
            
        with transaction.atomic():
            try:
                op.finalize_customs()
                op.save()
                self.notificar_senior(op, "marcado como lista para envío")
                return Response({'status': 'ready_for_delivery'})
            except ValidationError as e:
                return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def mark_delivered(self, request, pk=None):
        op = self.get_object()
        try:
            op.mark_delivered()
            op.save()
            self.notificar_senior(op, "marcado como remitada")
            return Response({'status': 'delivered'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def close_operation(self, request, pk=None):
        op = self.get_object()
        try:
            op.close()
            op.save()
            self.notificar_senior(op, "cerrado")
            return Response({'status': 'closed'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'], url_path='upload_packing')
    def upload_packing(self, request, pk=None):
        op = self.get_object()
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        op.packing_list_file = request.FILES['file']
        op.save()
        return Response({'status': 'ok'})

    @action(detail=True, methods=['post'], url_path='upload_remito')
    def upload_remito(self, request, pk=None):
        op = self.get_object()
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        op.remito_file = request.FILES['file']
        op.save()
        return Response({'status': 'ok'})

    @action(detail=True, methods=['post'], url_path='upload_rancho')
    def upload_rancho(self, request, pk=None):
        op = self.get_object()
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        op.rancho_file = request.FILES['file']
        op.save()
        return Response({'status': 'ok'})

    @action(detail=True, methods=['post'], url_path='upload_solicitud_particular')
    def upload_solicitud_particular(self, request, pk=None):
        op = self.get_object()
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
        op.solicitud_particular_file = request.FILES['file']
        op.save()
        return Response({'status': 'ok'})

    @action(detail=True, methods=['get'])
    def generate_solicitud_particular_pdf(self, request, pk=None):
        op = self.get_object()
        from apps.operaciones.services_pdf import generar_solicitud_particular_pdf
        try:
            pdf_bytes = generar_solicitud_particular_pdf(op)
            from django.http import HttpResponse
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="Solicitud_Particular_OP{op.id}.pdf"'
            return response
        except Exception as e:
            logger.exception("Error generando PDF de solicitud particular")
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['get'], url_path='generate_remito_docx')
    def generate_remito_docx(self, request, pk=None):
        op = self.get_object()
        from apps.operaciones.services_docx import generar_remito_docx
        try:
            docx_bytes = generar_remito_docx(op)
            from django.http import HttpResponse
            response = HttpResponse(docx_bytes, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = f'attachment; filename="Remito_OP{op.id}.docx"'
            return response
        except Exception as e:
            logger.exception("Error generando Remito DOCX")
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['get'], url_path='packing_list_json')
    def packing_list_json(self, request, pk=None):
        op = self.get_object()
        from apps.inventario.models import Articulo
        productos = []
        total_price = 0
        total_weight = 0
        for detalle in op.detalles.all():
            try:
                articulo = Articulo.objects.get(id=detalle.articulo_id)
                subtotal = detalle.cantidad * float(detalle.precio_unitario)
                total_price += subtotal
                unit_weight = float(articulo.peso_kg)
                total_weight += detalle.cantidad * unit_weight
                productos.append({
                    'name': articulo.nombre,
                    'quantity': detalle.cantidad,
                    'presentation': articulo.presentacion,
                    'unit_weight': unit_weight,
                    'total_weight': detalle.cantidad * unit_weight,
                    'unit_price': float(detalle.precio_unitario),
                    'subtotal': subtotal,
                })
            except Articulo.DoesNotExist:
                productos.append({
                    'name': f"Artículo ID {detalle.articulo_id} (no existe)",
                    'quantity': detalle.cantidad,
                    'presentation': '',
                    'unit_weight': 0,
                    'total_weight': 0,
                    'unit_price': float(detalle.precio_unitario),
                    'subtotal': detalle.cantidad * float(detalle.precio_unitario),
                })
        data = {
            'operation_id': op.id,
            'client': op.cliente.name,
            'ship': op.ship.name,
            'port': op.port.name,
            'eta': op.eta,
            'products': productos,
            'total_price': total_price,
            'total_weight': total_weight,
        }
        return Response(data)

    @action(detail=True, methods=['get'])
    def verificar_stock(self, request, pk=None):
        operacion = self.get_object()
        ok, errores = operacion.verificar_stock()

        def get_formula_shortage(articulo, shortage_qty):
            from apps.produccion.models import FormulaBOM
            from apps.inventario.models import Articulo
            from decimal import Decimal
            formula = FormulaBOM.objects.filter(articulo_final_id=articulo.id, activa=True).first()
            if not formula:
                return []
            
            shortage_qty_dec = Decimal(str(shortage_qty))
            shortage_list = []
            for comp in formula.componentes.all():
                try:
                    ing = Articulo.objects.get(id=comp.insumo_id)
                    necesario_ing = shortage_qty_dec * comp.cantidad_requerida
                    if ing.stock_actual < necesario_ing:
                        falta_ing = necesario_ing - ing.stock_actual
                        shortage_list.append({
                            'insumo_id': ing.id,
                            'nombre': ing.nombre,
                            'presentacion': ing.presentacion,
                            'unidad': ing.unidad or 'L',
                            'necesario': float(necesario_ing),
                            'disponible': float(ing.stock_actual),
                            'falta': float(falta_ing)
                        })
                except Articulo.DoesNotExist:
                    pass
            return shortage_list

        # Enriquecer errores
        from apps.inventario.models import Articulo
        for err in errores:
            articulo_id = err.get('articulo_id')
            if articulo_id:
                try:
                    articulo = Articulo.objects.get(id=articulo_id)
                    err['unidad'] = articulo.unidad or 'L'
                    if articulo.categoria == 'quimicos':
                        shortage_qty = err['necesario'] - err['disponible']
                        err['formula_shortage'] = get_formula_shortage(articulo, shortage_qty)
                except Articulo.DoesNotExist:
                    pass

        detalles_data = []
        for detalle in operacion.detalles.all():
            try:
                articulo = Articulo.objects.get(id=detalle.articulo_id)
                suficiente = True if not articulo.controlar_stock else (articulo.stock_actual >= detalle.cantidad)
                
                formula_shortage = []
                if not suficiente and articulo.categoria == 'quimicos':
                    shortage_qty = detalle.cantidad - articulo.stock_actual
                    formula_shortage = get_formula_shortage(articulo, shortage_qty)

                detalles_data.append({
                    'id': detalle.id,
                    'articulo_id': detalle.articulo_id,
                    'nombre': articulo.nombre,
                    'presentacion': articulo.presentacion,
                    'unidad': articulo.unidad or 'L',
                    'cantidad_necesaria': detalle.cantidad,
                    'stock_actual': float(articulo.stock_actual),
                    'suficiente': suficiente,
                    'controlar_stock': articulo.controlar_stock,
                    'error': None,
                    'formula_shortage': formula_shortage
                })
            except Articulo.DoesNotExist:
                detalles_data.append({
                    'id': detalle.id,
                    'articulo_id': detalle.articulo_id,
                    'nombre': f"ID {detalle.articulo_id}",
                    'presentacion': "",
                    'cantidad_necesaria': detalle.cantidad,
                    'stock_actual': 0,
                    'suficiente': False,
                    'controlar_stock': True,
                    'error': "Producto no existe en inventario",
                    'formula_shortage': []
                })

        return Response({
            'operacion_id': operacion.id,
            'todo_suficiente': ok,
            'detalles': detalles_data,
            'errores': errores
        })

    @action(detail=False, methods=['get'], url_path='auto_complete_imo')
    def auto_complete_imo(self, request):
        imo = request.query_params.get('imo')

        if not imo or not imo.isdigit() or len(imo) != 7:
            return Response({"error": "Se requiere un IMO válido de 7 dígitos"}, status=400)

        ship, scraped_data = get_or_create_ship_from_imo(imo)

        if not scraped_data:
            return Response({"error": "No se pudo obtener información del buque."}, status=404)

        port = None
        port_id = None
        port_name = scraped_data.get('destino')

        if port_name:
            port = get_or_create_port_from_name(port_name)
            port_id = port.id if port else None

        return Response({
            "ship_id": ship.id,
            "ship_name": ship.name,
            "flag": ship.flag,
            "imo": ship.imo,
            "eta": scraped_data.get('eta'),
            "eta_raw": scraped_data.get('eta_raw'),
            "port_id": port_id,
            "port_name": port_name,
        })

    @action(detail=True, methods=['get'], url_path='packing_list_excel')
    def packing_list_excel(self, request, pk=None):
        try:
            op = self.get_object()
            from apps.inventario.models import Articulo

            proveedor_seleccionado = request.query_params.get('proveedor', '')
            pais_destino_seleccionado = request.query_params.get('pais_destino', 'argentina')

            if pais_destino_seleccionado == 'argentina':
                pais_destino_texto = 'Argentina'
            else:
                pais_destino_texto = op.ship.flag if op.ship and op.ship.flag else 'Extranjero'

            if proveedor_seleccionado == 'PROIOS SA':
                proveedor_texto = 'PROIOS SA'
                cuit_texto = '30-63661723-3'
            elif proveedor_seleccionado == 'PROIOS SALVAGE SA':
                proveedor_texto = 'PROIOS SALVAGE SA'
                cuit_texto = '33-71087653-9'
            else:
                proveedor_texto = ''
                cuit_texto = ''

            productos = []
            total_price = 0.0
            total_weight_neto = 0.0
            total_weight_bruto = 0.0
            total_qty = 0

            for detalle in op.detalles.all():
                try:
                    articulo = Articulo.objects.get(id=detalle.articulo_id)
                    cantidad = detalle.cantidad
                    precio = float(detalle.precio_unitario) if detalle.precio_unitario else 0.0
                    subtotal = cantidad * precio
                    total_price += subtotal
                    total_qty += cantidad
                    peso_neto = float(articulo.peso_kg) if articulo.peso_kg is not None else 0.0
                    peso_bruto = peso_neto * 1.1
                    total_weight_neto += cantidad * peso_neto
                    total_weight_bruto += cantidad * peso_bruto
                    productos.append({
                        'descripcion': articulo.nombre,
                        'qty': cantidad,
                        'peso_neto': peso_neto,
                        'peso_bruto': peso_bruto,
                        'un_vta': articulo.presentacion or 'Kg',
                        'fob_unitario': precio,
                        'fob_total': subtotal,
                    })
                except Articulo.DoesNotExist:
                    logger.warning(f"Artículo {detalle.articulo_id} no existe en operación {op.id}")
                    continue

            wb = Workbook()
            ws = wb.active
            ws.title = "PACKING LIST"

            header_font = Font(bold=True, name='Arial')
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            gray_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            encabezados = [
                ("PROVEEDOR:", proveedor_texto),
                ("CUIT:", cuit_texto),
                ("PAÍS DE DESTINO DE LA FACTURA:", pais_destino_texto),
                ("BANDERA:", op.ship.flag if op.ship else ""),
                ("EMPRESA A FACTURAR:", op.cliente.name if op.cliente else ""),
                ("BUQUE:", op.ship.name if op.ship else "")
            ]
            row = 1
            for label, value in encabezados:
                ws.cell(row=row, column=1, value=label).font = header_font
                ws.cell(row=row, column=2, value=value)
                row += 1

            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            items_cell = ws.cell(row=row, column=1, value="ITEMS")
            items_cell.font = Font(bold=True, size=12)
            items_cell.alignment = center_align
            items_cell.fill = green_fill

            headers = ["DESCRIPCION", "QTY", "PESO NETO", "PESO BRUTO", "UN. VTA", "FOB UNITARIO", "FOB TOTAL"]
            row += 1
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = gray_header_fill

            for prod in productos:
                row += 1
                ws.cell(row=row, column=1, value=prod['descripcion'])
                ws.cell(row=row, column=2, value=prod['qty'])
                ws.cell(row=row, column=3, value=round(prod['peso_neto'], 2))
                ws.cell(row=row, column=4, value=round(prod['peso_bruto'], 2))
                ws.cell(row=row, column=5, value=prod['un_vta'])
                ws.cell(row=row, column=6, value=round(prod['fob_unitario'], 2))
                ws.cell(row=row, column=7, value=round(prod['fob_total'], 2))

            row += 1
            ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_qty)
            ws.cell(row=row, column=3, value=round(total_weight_neto, 2))
            ws.cell(row=row, column=4, value=round(total_weight_bruto, 2))
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value=f"USD {round(total_price, 2)}")
            ws.cell(row=row, column=7, value=f"USD {round(total_price, 2)}")

            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = orange_fill
                ws.cell(row=row, column=col).alignment = center_align

            for r in range(1, row+1):
                for c in range(1, 8):
                    ws.cell(row=r, column=c).border = thin_border

            column_widths = [40, 8, 12, 12, 10, 15, 15]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="packing_list_{op.id}.xlsx"'
            return response

        except Exception as e:
            logger.exception("Error generando packing list Excel")
            return Response({"error": str(e)}, status=500)

    @action(detail=True, methods=['post'])
    def request_review(self, request, pk=None):
        op = self.get_object()
        mensaje = request.data.get('mensaje_revision', '')
        
        op.estado_revision = Operacion.ESTADO_REVISION_PENDING
        op.aprobacion_requerida_owner = True
        if mensaje:
            op.mensaje_revision = mensaje
        op.save()
        return Response({'status': 'review_requested'})

    @action(detail=True, methods=['post'])
    def resolve_review(self, request, pk=None):
        op = self.get_object()
        user = request.user
        
        if user.role != User.Role.OWNER:
            return Response({'error': 'Solo el Owner puede resolver revisiones'}, status=403)
            
        action = request.data.get('action')
        mensaje = request.data.get('mensaje_revision', '')
        
        if action == 'approve':
            op.estado_revision = Operacion.ESTADO_REVISION_APPROVED
            op.aprobacion_requerida_owner = False
        elif action == 'reject':
            op.estado_revision = Operacion.ESTADO_REVISION_REJECTED
            op.aprobacion_requerida_owner = False
        else:
            return Response({'error': 'Acción inválida'}, status=400)
            
        if mensaje:
            op.mensaje_revision = mensaje
            
        op.save()
        return Response({'status': 'review_resolved', 'action': action})

    @action(detail=True, methods=['post'], url_path='documentos')
    def upload_documento(self, request, pk=None):
        """Subir un documento adicional a la operación"""
        operacion = self.get_object()
        tipo = request.data.get('tipo')
        archivo = request.FILES.get('archivo')
        nombre_personalizado = request.data.get('nombre_personalizado', '')
        descripcion = request.data.get('descripcion', '')

        if not tipo or not archivo:
            return Response({'error': 'Faltan campos obligatorios (tipo, archivo)'}, status=400)

        if tipo == 'otros' and not nombre_personalizado:
            return Response({'error': 'Para tipo "Otros" debe especificar un nombre personalizado'}, status=400)

        documento = DocumentoAdjunto.objects.create(
            operacion=operacion,
            tipo=tipo,
            nombre_personalizado=nombre_personalizado if tipo == 'otros' else '',
            archivo=archivo,
            descripcion=descripcion,
            subido_por=request.user if request.user.is_authenticated else None
        )
        return Response(DocumentoAdjuntoSerializer(documento).data, status=201)

    @action(detail=True, methods=['delete'], url_path='documentos/(?P<doc_id>[^/.]+)')
    def delete_documento(self, request, pk=None, doc_id=None):
        """Eliminar un documento adicional"""
        operacion = self.get_object()
        try:
            documento = operacion.documentos_adjuntos.get(id=doc_id)
            documento.archivo.delete()
            documento.delete()
            return Response({'status': 'deleted'})
        except DocumentoAdjunto.DoesNotExist:
            return Response({'error': 'Documento no encontrado'}, status=404)

    @action(detail=True, methods=['post'], url_path='descargar_zip')
    def descargar_zip(self, request, pk=None):
        """Descargar múltiples archivos de la operación agrupados en un archivo ZIP"""
        operacion = self.get_object()
        documento_ids = request.data.get('documento_ids', [])
        adjunto_ids = request.data.get('adjunto_ids', [])

        import zipfile
        import io
        from apps.correos.models import EmailAttachment

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            if documento_ids:
                docs = operacion.documentos_adjuntos.filter(id__in=documento_ids)
                for doc in docs:
                    if doc.archivo:
                        filename = doc.archivo.name.split('/')[-1]
                        try:
                            doc.archivo.open('rb')
                            zip_file.writestr(f"documentos/{filename}", doc.archivo.read())
                            doc.archivo.close()
                        except Exception as e:
                            logger.error(f"Error agregando documento {doc.id} al zip: {e}")

            if adjunto_ids:
                adjuntos = EmailAttachment.objects.filter(id__in=adjunto_ids, email__operacion=operacion)
                for adj in adjuntos:
                    if adj.file:
                        filename = adj.filename or adj.file.name.split('/')[-1]
                        try:
                            adj.file.open('rb')
                            zip_file.writestr(f"adjuntos_correos/{filename}", adj.file.read())
                            adj.file.close()
                        except Exception as e:
                            logger.error(f"Error agregando adjunto {adj.id} al zip: {e}")

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/x-zip-compressed')
        response['Content-Disposition'] = f'attachment; filename="operacion_{operacion.id}_archivos.zip"'
        return response

    @action(detail=True, methods=['post'], url_path='generar_ordenes_produccion')
    def generar_ordenes_produccion(self, request, pk=None):
        if request.user.role == User.Role.OPERARIO:
            return Response({'error': 'No autorizado para generar órdenes de producción'}, status=status.HTTP_403_FORBIDDEN)
            
        op = self.get_object()
        from apps.produccion.models import FormulaBOM, OrdenFabricacion
        from apps.inventario.models import Articulo
        
        ordenes_creadas = []
        with transaction.atomic():
            for detalle in op.detalles.all():
                try:
                    articulo = Articulo.objects.select_for_update().get(id=detalle.articulo_id)
                except Articulo.DoesNotExist:
                    continue
                
                formula = FormulaBOM.objects.filter(articulo_final_id=articulo.id, activa=True).first()
                if not formula:
                    continue
                
                shortage = detalle.cantidad - articulo.stock_actual
                if shortage > 0:
                    if not OrdenFabricacion.objects.filter(operacion_id=op.id, formula=formula, completada=False).exists():
                        orden = OrdenFabricacion.objects.create(
                            operacion_id=op.id,
                            formula=formula,
                            cantidad_a_producir=shortage,
                            completada=False
                        )
                        ordenes_creadas.append(orden)
        
        from apps.produccion.api.serializers import OrdenFabricacionSerializer
        return Response({
            'status': 'success',
            'created_count': len(ordenes_creadas),
            'ordenes': OrdenFabricacionSerializer(ordenes_creadas, many=True).data
        })

    @action(detail=True, methods=['get'], url_path='ordenes_produccion')
    def ordenes_produccion(self, request, pk=None):
        if request.user.role == User.Role.OPERARIO:
            return Response({'error': 'No autorizado para ver órdenes de producción'}, status=status.HTTP_403_FORBIDDEN)
            
        op = self.get_object()
        from apps.produccion.models import OrdenFabricacion
        from apps.produccion.api.serializers import OrdenFabricacionSerializer
        
        ordenes = OrdenFabricacion.objects.filter(operacion_id=op.id)
        return Response(OrdenFabricacionSerializer(ordenes, many=True).data)

    # ========== NUEVO ENDPOINT PARA OPERARIOS ==========
    @action(detail=False, methods=['get'], url_path='tareas-produccion')
    def tareas_produccion(self, request):
        """
        Devuelve las órdenes de fabricación pendientes
        SOLO de las operaciones donde el operario está asignado.
        """
        user = request.user
        print(f"🔍 [BACKEND tareas-produccion] Usuario: {user.username} (role: {user.role})")
        
        if user.role != User.Role.OPERARIO:
            return Response({'error': 'Solo para operarios'}, status=403)
        
        # Obtener IDs de operaciones donde el usuario está asignado como operario
        operaciones_ids = Operacion.objects.filter(
            operarios_usuarios_asignados=user
        ).values_list('id', flat=True)
        
        print(f"📋 [tareas-produccion] Operaciones asignadas IDs: {list(operaciones_ids)}")
        
        if not operaciones_ids:
            return Response([])
        
        ordenes = OrdenFabricacion.objects.filter(
            completada=False,
            operacion_id__in=operaciones_ids
        ).order_by('-fecha_solicitud')
        
        print(f"📦 [tareas-produccion] Órdenes pendientes encontradas: {ordenes.count()}")
        
        resultados = []
        for orden in ordenes:
            try:
                operacion = Operacion.objects.get(id=orden.operacion_id)
                op_data = {
                    'id': operacion.id,
                    'nombre': operacion.nombre or f"OP-{operacion.id:05d}",
                    'ship_name': operacion.ship.name if operacion.ship else '',
                    'port_name': operacion.port.name if operacion.port else '',
                    'estado': operacion.get_estado_display(),
                    'remito_file': operacion.remito_file.url if operacion.remito_file else None,
                    'packing_list_file': operacion.packing_list_file.url if operacion.packing_list_file else None,
                    'eta': operacion.eta,
                }
            except Operacion.DoesNotExist:
                op_data = {
                    'id': orden.operacion_id,
                    'nombre': f"Operación #{orden.operacion_id} (no encontrada)",
                    'ship_name': '',
                    'port_name': '',
                    'estado': '',
                    'remito_file': None,
                    'packing_list_file': None,
                    'eta': None,
                }
            
            resultados.append({
                'orden_id': orden.id,
                'operacion': op_data,
                'producto_nombre': orden.formula.articulo_final_nombre,
                'cantidad_a_producir': float(orden.cantidad_a_producir),
                'fecha_solicitud': orden.fecha_solicitud,
                'completada': orden.completada,
            })
        
        return Response(resultados)

    # ========== ENDPOINT AUXILIAR PARA ASIGNAR OPERARIOS ==========
    @action(detail=True, methods=['post'], url_path='asignar-operarios')
    def asignar_operarios(self, request, pk=None):
        """
        Endpoint específico para asignar operarios (usuarios con rol OPERARIO) a una operación.
        Body: {"operarios_usuarios_id": [1,2,3]}
        """
        operacion = self.get_object()
        operarios_ids = request.data.get('operarios_usuarios_id', [])
        
        if not isinstance(operarios_ids, list):
            return Response({'error': 'operarios_usuarios_id debe ser una lista de IDs'}, status=400)
        
        usuarios = User.objects.filter(id__in=operarios_ids)
        for u in usuarios:
            if u.role != User.Role.OPERARIO:
                return Response({'error': f'El usuario {u.username} no tiene rol OPERARIO'}, status=400)
        
        operacion.operarios_usuarios_asignados.set(operarios_ids)
        operacion.save()
        
        return Response({
            'status': 'ok',
            'operarios_usuarios_asignados': list(operacion.operarios_usuarios_asignados.values_list('id', flat=True))
        })


    # ========== SERVICIOS (FSM y Generación de PDF) ==========
    
    @action(detail=True, methods=['post'])
    def cotizar_servicio(self, request, pk=None):
        op = self.get_object()
        try:
            with transaction.atomic():
                op.cotizar_servicio()
                op.save()
            return Response({'status': 'cotizado'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def tramitar_permisos_pna(self, request, pk=None):
        op = self.get_object()
        try:
            with transaction.atomic():
                op.tramitar_permisos_pna()
                op.save()
            return Response({'status': 'permisos_gestionados'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def iniciar_ejecucion_servicio(self, request, pk=None):
        op = self.get_object()
        try:
            with transaction.atomic():
                op.iniciar_ejecucion_servicio()
                op.save()
            return Response({'status': 'en_ejecucion'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def finalizar_servicio_reporte(self, request, pk=None):
        op = self.get_object()
        try:
            with transaction.atomic():
                op.finalizar_servicio_reporte()
                op.save()
            return Response({'status': 'reporte_firmado'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['post'])
    def close_servicio(self, request, pk=None):
        op = self.get_object()
        try:
            with transaction.atomic():
                op.close_servicio()
                op.save()
            return Response({'status': 'closed_servicio'})
        except ValidationError as e:
            return Response({'error': str(e)}, status=400)

    @action(detail=True, methods=['get'])
    def generate_cotizacion_servicio(self, request, pk=None):
        op = self.get_object()
        from apps.operaciones.services_docx import generar_cotizacion_servicio_docx
        
        # Obtener parametros para el PDF
        params = {
            'mobilization': request.query_params.get('mobilization', '2-3 days after order confirmation.'),
            'execution': request.query_params.get('execution', '1 day.'),
            'bank_charges': request.query_params.get('bank_charges', 'USD 50.00.'),
            'taxes': request.query_params.get('taxes', '(ISS 3%): will be applied to the total invoice value accordingly.'),
            'payment_terms': request.query_params.get('payment_terms', 'To Be Negotiated.'),
            'transport': request.query_params.get('transport', 'Transportation of the rotor from the vessel to Barcarena/Belém and back (round trip) is to be arranged by and under the responsibility of the client/agents.'),
        }

        try:
            docx_content = generar_cotizacion_servicio_docx(op, request.user, params)
            response = HttpResponse(
                docx_content, 
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = f'attachment; filename="cotizacion_servicio_OP{op.id}.docx"'
            return response
        except Exception as e:
            logger.exception("Error generando cotizacion de servicio")
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['get'])
    def generate_cotizacion_docx(self, request, pk=None):
        op = self.get_object()
        offer_validity = request.query_params.get('offer_validity', '15 days')
        payment_terms = request.query_params.get('payment_terms', '30 days from invoice date')
        warranty = request.query_params.get('warranty', 'As per manufacturer')

        from apps.operaciones.services_docx import generar_cotizacion_docx
        try:
            docx_content = generar_cotizacion_docx(op, offer_validity, payment_terms, warranty)
            response = HttpResponse(
                docx_content, 
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = f'attachment; filename="cotizacion_OP{op.id}.docx"'
            return response
        except Exception as e:
            logger.exception("Error generando cotizacion docx")
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['get'])
    def generate_permiso_pna(self, request, pk=None):
        op = self.get_object()
        tipo_trabajo = request.query_params.get('tipo', 'frio')
        from apps.operaciones.services_pdf import generar_permiso_pna_pdf
        try:
            pdf_content = generar_permiso_pna_pdf(op, tipo_trabajo)
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="permiso_pna_OP{op.id}.pdf"'
            return response
        except Exception as e:
            logger.exception("Error generando permiso PNA")
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['get'])
    def generate_solicitud_particular(self, request, pk=None):
        op = self.get_object()
        from apps.operaciones.services_pdf import generar_solicitud_particular_pdf
        try:
            pdf_content = generar_solicitud_particular_pdf(op)
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="solicitud_particular_OP{op.id}.pdf"'
            return response
        except Exception as e:
            logger.exception("Error generando solicitud particular")
            return Response({'error': str(e)}, status=500)

    @action(detail=True, methods=['get'])
    def generate_reporte_servicio(self, request, pk=None):
        op = self.get_object()
        from apps.operaciones.services_pdf import generar_reporte_servicio_pdf
        try:
            pdf_content = generar_reporte_servicio_pdf(op)
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_servicio_OP{op.id}.pdf"'
            return response
        except Exception as e:
            logger.exception("Error generando reporte servicio")
            return Response({'error': str(e)}, status=500)


class AgendaEventViewSet(viewsets.ModelViewSet):
    serializer_class = AgendaEventSerializer

    def get_queryset(self):
        user = self.request.user
        qs = AgendaEvent.objects.all().select_related('created_by', 'assigned_to')

        if user.role in [User.Role.OWNER, User.Role.CONTABLE]:
            user_id = self.request.query_params.get('user_id')
            if user_id:
                qs = qs.filter(assigned_to_id=user_id)
            return qs

        return qs.filter(Q(assigned_to=user) | Q(created_by=user)).distinct()

    def perform_create(self, serializer):
        assigned_to = serializer.validated_data.get('assigned_to')
        user = self.request.user
        
        if user.role not in [User.Role.OWNER, User.Role.CONTABLE] and assigned_to != user:
            raise ValidationError("No tiene permiso para asignar eventos a otros usuarios.")
            
        instance = serializer.save(created_by=user)
        logger.info(f"Nuevo Evento de Agenda creado por {user.username}: {instance.title} (Asignado a: {instance.assigned_to.username})")

    def perform_update(self, serializer):
        instance = serializer.save()
        logger.info(f"Evento de Agenda actualizado por {self.request.user.username}: {instance.title}")

    def perform_destroy(self, instance):
        logger.info(f"Evento de Agenda eliminado por {self.request.user.username}: {instance.title}")
        instance.delete()

from rest_framework.views import APIView
from apps.usuarios.permissions import IsLocalIP

class TvDashboardView(APIView):
    permission_classes = [IsLocalIP]
    authentication_classes = []

    def get(self, request, *args, **kwargs):
        # Operaciones que no esten entregadas ni canceladas
        qs = Operacion.objects.exclude(estado__in=[
            Operacion.ESTADO_ENTREGADA,
            Operacion.ESTADO_CANCELADA
        ]).select_related('cliente', 'ship')
        
        data = []
        for op in qs:
            operadores = op.operadores_asignados.all()
            operadores_nombres = [u.first_name or u.username for u in operadores]
            
            data.append({
                'id': op.id,
                'nombre': op.nombre or f"OP-{op.id:05d}",
                'cliente': op.cliente.name if op.cliente else 'N/A',
                'buque': op.ship.name if op.ship else 'N/A',
                'eta': op.eta,
                'estado': op.get_estado_display(),
                'estado_raw': op.estado,
                'tipo_operacion': op.get_tipo_operacion_display(),
                'operadores': operadores_nombres
            })
            
        return Response({'operaciones': data})