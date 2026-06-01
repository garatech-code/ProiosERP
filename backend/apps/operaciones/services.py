from django.core.exceptions import PermissionDenied
from .models import Operacion, Ship, Port
from apps.usuarios.models import User
import re
import logging
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from django.utils import timezone

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------
# Servicio de operaciones (lógica de negocio)
# ----------------------------------------------------------------------
class OperacionService:
    """
    Capa de Servicios para la lógica de negocio de Operaciones.
    Centraliza la Máquina de Estados y el control de transiciones basado en Roles.
    """

    @staticmethod
    def _verificar_y_forzar_borrador(operacion: Operacion, usuario: User) -> bool:
        """
        Interpreta si el cambio propuesto por el usuario requiere pasar por aprobación.
        Retorna True si debe quedar PENDIENTE DE APROBACIÓN, False si viaja directo al estado.
        """
        if usuario.role == User.Role.OWNER:
            return False
            
        if usuario.role in [User.Role.OPERADOR, User.Role.CONTABLE]:
            if usuario.requires_owner_review:
                return True
            return False

        # Los operarios no pueden modificar estados operativos, eleva error de seguridad. 
        raise PermissionDenied("Los operarios de planta no pueden alterar el ciclo de vida documental.")

    @staticmethod
    def procesar_presupuesto(operacion: Operacion, usuario: User):
        """Intenta lanzar la operacion a Presupuestada"""
        if OperacionService._verificar_y_forzar_borrador(operacion, usuario):
            operacion.estado_revision = 'pending'
        else:
            operacion.presupuestar()
        
        operacion.save()
        # Aquí se emitirá el evento Celery en el futuro

    @staticmethod
    def iniciar_produccion(operacion: Operacion, usuario: User):
        if OperacionService._verificar_y_forzar_borrador(operacion, usuario):
            operacion.estado_revision = 'pending'
        else:
            operacion.iniciar_produccion()
        operacion.save()

    @staticmethod
    def aprobar_cambios_por_owner(operacion: Operacion, nuevo_estado: str, owner: User):
        """
        Método exclusivo para el Owner: Fuerza la transición de un borrador hacia adelante.
        """
        if owner.role != User.Role.OWNER:
            raise PermissionDenied("Solo el Owner puede aprobar transiciones bloqueadas.")

        if operacion.estado_revision != 'pending':
            raise ValueError("La operación no está pendiente de revisión.")

        # Bypass de FSM estricto mediante override forzado del estado aprobado por el Owner
        operacion.estado_revision = 'approved'
        operacion.save()


# ----------------------------------------------------------------------
# Servicios de scraping y autocompletado por IMO
# ----------------------------------------------------------------------
def scrape_vessel_info(imo):
    """
    Extrae nombre, bandera, destino (puerto) y ETA de VesselFinder.
    Retorna un dict con los datos o None si hay error.
    - La lógica de ETA se mantiene exactamente como la versión que funciona bien.
    - El resto (nombre, bandera, destino) se ha mejorado con múltiples estrategias y limpieza.
    """
    import re
    from datetime import datetime
    import requests
    from bs4 import BeautifulSoup
    from django.utils import timezone

    url = f"https://www.vesselfinder.com/vessels/details/{imo}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"Error scraping IMO {imo} desde VesselFinder: {e}")
        return None

    soup = BeautifulSoup(response.text, "html.parser")
    text = soup.get_text()

    # ---------- NOMBRE (mejorado) ----------
    nombre = None
    h1 = soup.find("h1", class_="title")
    if h1:
        nombre = h1.get_text(strip=True)
    if not nombre:
        h1_alt = soup.find("h1")
        if h1_alt:
            nombre = h1_alt.get_text(strip=True)
    # fallback: buscar en el párrafo text2
    if not nombre:
        paragraph = soup.find("p", class_="text2")
        if paragraph:
            match = re.search(r"The vessel\s+([A-Za-z0-9\s]+)\s+\(IMO", paragraph.get_text())
            if match:
                nombre = match.group(1).strip()

    # ---------- TIPO (del subtítulo h2) ----------
    vessel_type = None
    h2 = soup.find("h2", class_="vst")
    if h2:
        vessel_type = h2.get_text(strip=True).split(",")[0].strip()

    # ---------- BANDERA (mejorada: prioriza nombre completo desde título o mapeo) ----------
    bandera = None
    # Mapeo de códigos a nombres (ampliado)
    flag_map = {
        "mt": "Malta", "ar": "Argentina", "br": "Brazil", "sg": "Singapore",
        "hk": "Hong Kong", "mh": "Marshall Islands", "bg": "Bulgaria",
        "bb": "Barbados", "lr": "Liberia", "pa": "Panama", "uy": "Uruguay",
        "us": "United States", "gb": "United Kingdom", "es": "Spain", "fr": "France",
        "de": "Germany", "it": "Italy", "nl": "Netherlands", "tr": "Turkey",
        "ru": "Russia", "cn": "China", "jp": "Japan", "kr": "South Korea",
    }
    # Método 1: div title-flag-icon (tiene style con la bandera)
    flag_div = soup.find("div", class_="title-flag-icon")
    if flag_div:
        # Primero ver si tiene atributo 'title' (nombre completo)
        if flag_div.get("title"):
            bandera = flag_div["title"]
        elif flag_div.get("style"):
            match = re.search(r'flags/4x3/([a-z]+)\.svg', flag_div["style"])
            if match:
                code = match.group(1).lower()
                bandera = flag_map.get(code, code.upper())
    # Método 2: párrafo text2 ("flag of Malta")
    if not bandera:
        paragraph = soup.find("p", class_="text2")
        if paragraph:
            text_p = paragraph.get_text()
            match = re.search(r"flag of\s+([A-Za-z\s]+)", text_p, re.IGNORECASE)
            if match:
                bandera = match.group(1).strip()
    # Método 3: tabla "AIS Flag"
    if not bandera:
        flag_row = soup.find("td", string=re.compile(r"AIS Flag", re.IGNORECASE))
        if flag_row:
            flag_td = flag_row.find_next_sibling("td")
            if flag_td:
                val = flag_td.get_text(strip=True)
                bandera = flag_map.get(val.lower(), val)
    # Método 4: tabla "Flag" general
    if not bandera:
        flag_row = soup.find("td", string=re.compile(r"Flag", re.IGNORECASE))
        if flag_row:
            flag_td = flag_row.find_next_sibling("td")
            if flag_td:
                val = flag_td.get_text(strip=True)
                bandera = flag_map.get(val.lower(), val)

    # ---------- DESTINO (mejorado: captura cualquier destino, con o sin país) ----------
    destino = None
    paragraph = soup.find("p", class_="text2")
    if paragraph:
        text_p = paragraph.get_text()
        # Captura desde "en route to" hasta "sailing" o punto o fin de línea
        match_dest = re.search(r"en route to\s+(?:(?:the\s+port\s+of\s+)?)(.+?)(?:\s+sailing|\.|$)", text_p, re.IGNORECASE)
        if match_dest:
            destino = match_dest.group(1).strip()
            # Limpiar comas finales y espacios
            destino = re.sub(r',\s*$', '', destino)
            # Normalizar espacios alrededor de comas
            destino = re.sub(r'\s*,\s*', ', ', destino)
    # Método 2: tabla Voyage Data -> "Destination"
    if not destino:
        dest_label = soup.find("td", string=re.compile(r"Destination", re.IGNORECASE))
        if dest_label:
            dest_td = dest_label.find_next_sibling("td")
            if dest_td:
                destino = dest_td.get_text(strip=True)
    # Método 3: enlace con clase "_npNa"
    if not destino:
        dest_link = soup.find("a", class_="_npNa")
        if dest_link:
            destino = dest_link.get_text(strip=True)
    # Método 4: div "_3-Yih"
    if not destino:
        dest_div = soup.find("div", class_="_3-Yih")
        if dest_div:
            destino = dest_div.get_text(strip=True)

    # ---------- ETA (EXACTAMENTE IGUAL A TU CÓDIGO QUE FUNCIONA) ----------
    eta_str = None
    eta_iso = None
    # Método 1: del párrafo "expected to arrive there on Apr 8, 14:00"
    if paragraph:
        text_p = paragraph.get_text()
        match_eta = re.search(r"arrive there on\s+([A-Za-z]+\s+\d{1,2},\s+\d{2}:\d{2})", text_p)
        if match_eta:
            eta_str = match_eta.group(1).strip()
    # Método 2: de la tabla Voyage Data -> "Predicted ETA"
    if not eta_str:
        eta_label = soup.find("td", string=re.compile(r"Predicted ETA", re.IGNORECASE))
        if eta_label:
            eta_td = eta_label.find_next_sibling("td")
            if eta_td:
                eta_str = eta_td.get_text(strip=True)
                if eta_str == "-":
                    eta_str = None
    # Método 3: del elemento "_value" dentro de la sección de destino
    if not eta_str:
        value_div = soup.find("div", class_="_value")
        if value_div:
            eta_text = value_div.get_text()
            match_eta = re.search(r"ETA:\s*([^)]+)\)", eta_text)
            if match_eta:
                eta_str = match_eta.group(1).strip()

    # Parsear ETA a ISO (mismo código que usas)
    if eta_str:
        logger.info(f"ETA raw para IMO {imo}: {eta_str}")
        eta_clean = eta_str.replace('.', '').strip()
        now = timezone.now()
        for fmt in ["%b %d, %Y %H:%M", "%B %d, %Y %H:%M", "%b %d, %H:%M", "%B %d, %H:%M"]:
            try:
                dt = datetime.strptime(eta_clean, fmt)
                if dt.year == 1900:
                    dt = dt.replace(year=now.year)
                eta_iso = timezone.make_aware(dt).isoformat()
                logger.info(f"ETA parseado: {eta_iso}")
                break
            except ValueError:
                continue
        if not eta_iso:
            try:
                from dateutil import parser
                dt = parser.parse(eta_str, fuzzy=True)
                eta_iso = timezone.make_aware(dt).isoformat()
            except:
                pass
        if not eta_iso:
            logger.warning(f"No se pudo parsear ETA para {imo}: {eta_str}")

    # ---------- DATOS ADICIONALES (velocidad, calado) ----------
    speed = None
    draught = None
    speed_label = soup.find("td", string=re.compile(r"Course / Speed", re.IGNORECASE))
    if speed_label:
        speed_td = speed_label.find_next_sibling("td")
        if speed_td:
            speed_text = speed_td.get_text()
            match_speed = re.search(r"([\d\.]+)\s*knots", speed_text)
            if match_speed:
                speed = match_speed.group(1)
    draught_label = soup.find("td", string=re.compile(r"Current draught", re.IGNORECASE))
    if draught_label:
        draught_td = draught_label.find_next_sibling("td")
        if draught_td:
            draught = draught_td.get_text(strip=True)

    # ---------- RETORNO ----------
    return {
        "imo": imo,
        "nombre": nombre,
        "bandera": bandera,
        "destino": destino,
        "eta": eta_iso,
        "eta_raw": eta_str,
        "tipo": vessel_type,
        "velocidad_nudos": speed,
        "calado_metros": draught,
    }  
    
def get_or_create_ship_from_imo(imo):
    """Busca o crea un Ship usando el IMO, y opcionalmente actualiza nombre/bandera desde scraping."""
    try:
        ship = Ship.objects.get(imo=imo)
    except Ship.DoesNotExist:
        ship = None
    
    data = scrape_vessel_info(imo)
    if data:
        defaults = {
            'name': data['nombre'] or f"Buque {imo}",
            'flag': data['bandera'] or 'Desconocida',
        }
        ship, created = Ship.objects.update_or_create(
            imo=imo,
            defaults=defaults
        )
        return ship, data
    else:
        if not ship:
            ship = Ship.objects.create(
                imo=imo,
                name=f"Buque {imo}",
                flag="Desconocida"
            )
        return ship, None


def get_or_create_port_from_name(port_name):
    """
    Busca o crea un Puerto por nombre (que puede incluir país, ej. 'San Lorenzo, Argentina').
    Si el nombre contiene una coma, intenta separar ciudad y país para llenar los campos respectivos.
    """
    if not port_name:
        return None

    # Limpiar nombre
    port_name = port_name.strip()
    
    # Intentar separar ciudad y país si hay coma
    if ',' in port_name:
        parts = port_name.split(',', 1)
        city = parts[0].strip()
        country = parts[1].strip()
    else:
        city = port_name
        country = "Desconocido"

    # Buscar por nombre exacto (ciudad + país) o solo ciudad? Para evitar duplicados, buscamos primero por el nombre completo
    port, created = Port.objects.get_or_create(
        name=port_name,
        defaults={'country': country}
    )
    # Si ya existía con ese nombre pero el país está vacío o incorrecto, actualizamos
    if not created and (port.country == "Desconocido" or port.country != country):
        port.country = country
        port.save(update_fields=['country'])
    
    return port


def generar_y_guardar_packing_list(op, proveedor="PROIOS SA", pais_destino="argentina"):
    """
    Genera el archivo Excel de Packing List para una operación y lo guarda en su campo packing_list_file.
    """
    from apps.inventario.models import Articulo
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.core.files.base import ContentFile
    import io

    if pais_destino == 'argentina':
        pais_destino_texto = 'Argentina'
    else:
        pais_destino_texto = op.ship.flag if op.ship and op.ship.flag else 'Extranjero'

    if proveedor == 'PROIOS SA':
        proveedor_texto = 'PROIOS SA'
        cuit_texto = '30-63661723-3'
    elif proveedor == 'PROIOS SALVAGE SA':
        proveedor_texto = 'PROIOS SALVAGE SA'
        cuit_texto = '33-71087653-9'
    else:
        proveedor_texto = proveedor or ''
        cuit_texto = ''

    productos = []
    total_price = 0.0
    total_weight_neto = 0.0
    total_weight_bruto = 0.0
    total_qty = 0

    for detalle in op.detalles.all():
        try:
            articulo = Articulo.objects.get(id=detalle.articulo_id)
            cantidad = detalle.cantidad
            precio = float(detalle.precio_unitario) if detalle.precio_unitario else 0.0
            subtotal = cantidad * precio
            total_price += subtotal
            total_qty += cantidad
            peso_neto = float(articulo.peso_kg) if articulo.peso_kg is not None else 0.0
            peso_bruto = peso_neto * 1.1
            total_weight_neto += cantidad * peso_neto
            total_weight_bruto += cantidad * peso_bruto
            productos.append({
                'descripcion': articulo.nombre,
                'qty': cantidad,
                'peso_neto': peso_neto,
                'peso_bruto': peso_bruto,
                'un_vta': articulo.presentacion or 'Kg',
                'fob_unitario': precio,
                'fob_total': subtotal,
            })
        except Articulo.DoesNotExist:
            continue

    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING LIST"

    header_font = Font(bold=True, name='Arial')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    gray_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    encabezados = [
        ("PROVEEDOR:", proveedor_texto),
        ("CUIT:", cuit_texto),
        ("PAÍS DE DESTINO DE LA FACTURA:", pais_destino_texto),
        ("BANDERA:", op.ship.flag if op.ship else ""),
        ("EMPRESA A FACTURAR:", op.cliente.name if op.cliente else ""),
        ("BUQUE:", op.ship.name if op.ship else "")
    ]
    row = 1
    for label, value in encabezados:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    items_cell = ws.cell(row=row, column=1, value="ITEMS")
    items_cell.font = Font(bold=True, size=12)
    items_cell.alignment = center_align
    items_cell.fill = green_fill

    headers = ["DESCRIPCION", "QTY", "PESO NETO", "PESO BRUTO", "UN. VTA", "FOB UNITARIO", "FOB TOTAL"]
    row += 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = gray_header_fill

    for prod in productos:
        row += 1
        ws.cell(row=row, column=1, value=prod['descripcion'])
        ws.cell(row=row, column=2, value=prod['qty'])
        ws.cell(row=row, column=3, value=round(prod['peso_neto'], 2))
        ws.cell(row=row, column=4, value=round(prod['peso_bruto'], 2))
        ws.cell(row=row, column=5, value=prod['un_vta'])
        ws.cell(row=row, column=6, value=round(prod['fob_unitario'], 2))
        ws.cell(row=row, column=7, value=round(prod['fob_total'], 2))

    row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_qty)
    ws.cell(row=row, column=3, value=round(total_weight_neto, 2))
    ws.cell(row=row, column=4, value=round(total_weight_bruto, 2))
    ws.cell(row=row, column=5, value="")
    ws.cell(row=row, column=6, value=f"USD {round(total_price, 2)}")
    ws.cell(row=row, column=7, value=f"USD {round(total_price, 2)}")

    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = orange_fill
        ws.cell(row=row, column=col).alignment = center_align

    for r in range(1, row+1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin_border

    column_widths = [40, 8, 12, 12, 10, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"packing_list_{op.id}.xlsx"
    op.packing_list_file.save(filename, ContentFile(output.getvalue()), save=True)